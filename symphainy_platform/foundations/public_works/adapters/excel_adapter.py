"""
Excel Processing Adapter - Layer 0

Raw technology client for Excel file parsing.
Uses openpyxl or pandas for parsing.

WHAT (Infrastructure): I provide Excel parsing capabilities
HOW (Adapter): I use openpyxl/pandas to parse Excel files
"""

import logging
from typing import Dict, Any, Optional
import io

logger = logging.getLogger(__name__)


class ExcelProcessingAdapter:
    """
    Adapter for Excel file parsing.
    
    Uses openpyxl or pandas for parsing XLSX/XLS files.
    """
    
    def __init__(self):
        """Initialize Excel Processing Adapter."""
        self.logger = logger
        self.pandas_available = False
        self.openpyxl_available = False
        
        # Try to import pandas
        try:
            import pandas as pd
            self.pandas = pd
            self.pandas_available = True
            self.logger.info("✅ Pandas available for Excel parsing")
        except ImportError:
            self.logger.warning("⚠️ Pandas not available, will try openpyxl")
        
        # Try to import openpyxl
        try:
            import openpyxl
            self.openpyxl = openpyxl
            self.openpyxl_available = True
            self.logger.info("✅ Openpyxl available for Excel parsing")
        except ImportError:
            self.logger.warning("⚠️ Openpyxl not available")
        
        if not self.pandas_available and not self.openpyxl_available:
            self.logger.warning("⚠️ Neither pandas nor openpyxl available - Excel parsing will be limited")
        
        self.logger.info("✅ Excel Processing Adapter initialized")
    
    async def parse_file(self, file_data: bytes, filename: str) -> Dict[str, Any]:
        """
        Parse Excel file from bytes.
        
        Args:
            file_data: Excel file content as bytes
            filename: Original filename (for logging)
        
        Returns:
            Dict with parsed data:
            {
                "success": bool,
                "sheets": List[Dict],  # List of sheet data
                "tables": List[Dict],  # List of tables (one per sheet)
                "metadata": Dict,
                "error": Optional[str]
            }
        """
        try:
            # Use pandas if available (preferred)
            if self.pandas_available:
                return await self._parse_with_pandas(file_data, filename)
            
            # Fallback to openpyxl
            if self.openpyxl_available:
                return await self._parse_with_openpyxl(file_data, filename)
            
            # No library available - return error
            return {
                "success": False,
                "error": "Neither pandas nor openpyxl available for Excel parsing",
                "sheets": [],
                "tables": [],
                "metadata": {}
            }
        
        except Exception as e:
            self.logger.error(f"❌ Excel parsing failed: {e}", exc_info=True)
            return {
                "success": False,
                "error": f"Excel parsing failed: {str(e)}",
                "sheets": [],
                "tables": [],
                "metadata": {}
            }
    
    async def _parse_with_pandas(self, file_data: bytes, filename: str) -> Dict[str, Any]:
        """Parse Excel file using pandas."""
        try:
            # Read Excel file from bytes
            excel_file = io.BytesIO(file_data)
            
            # Read all sheets
            excel_data = self.pandas.read_excel(excel_file, sheet_name=None, engine='openpyxl')
            
            sheets = []
            tables = []
            
            for sheet_name, df in excel_data.items():
                # Convert DataFrame to list of dictionaries
                rows = df.to_dict('records')
                
                # Get column names
                columns = list(df.columns)
                
                sheets.append({
                    "name": sheet_name,
                    "rows": rows,
                    "columns": columns,
                    "row_count": len(rows),
                    "column_count": len(columns)
                })
                
                tables.append({
                    "sheet_name": sheet_name,
                    "data": rows,
                    "columns": columns,
                    "row_count": len(rows)
                })
            
            metadata = {
                "type": "excel",
                "sheet_count": len(sheets),
                "filename": filename,
                "size": len(file_data)
            }
            
            return {
                "success": True,
                "sheets": sheets,
                "tables": tables,
                "metadata": metadata
            }
        
        except Exception as e:
            self.logger.error(f"❌ Pandas Excel parsing failed: {e}", exc_info=True)
            raise
    
    async def _parse_with_openpyxl(self, file_data: bytes, filename: str) -> Dict[str, Any]:
        """Parse Excel file using openpyxl."""
        try:
            # Read Excel file from bytes
            excel_file = io.BytesIO(file_data)
            workbook = self.openpyxl.load_workbook(excel_file, data_only=True)
            
            sheets = []
            tables = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Extract data
                rows = []
                headers = None
                
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    if row_idx == 0:
                        # First row is headers
                        headers = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(row)]
                    else:
                        # Data rows
                        row_dict = {}
                        for col_idx, cell_value in enumerate(row):
                            col_name = headers[col_idx] if headers and col_idx < len(headers) else f"Column_{col_idx+1}"
                            row_dict[col_name] = cell_value
                        rows.append(row_dict)
                
                sheets.append({
                    "name": sheet_name,
                    "rows": rows,
                    "columns": headers or [],
                    "row_count": len(rows),
                    "column_count": len(headers) if headers else 0
                })
                
                tables.append({
                    "sheet_name": sheet_name,
                    "data": rows,
                    "columns": headers or [],
                    "row_count": len(rows)
                })
            
            metadata = {
                "type": "excel",
                "sheet_count": len(sheets),
                "filename": filename,
                "size": len(file_data)
            }
            
            return {
                "success": True,
                "sheets": sheets,
                "tables": tables,
                "metadata": metadata
            }
        
        except Exception as e:
            self.logger.error(f"❌ Openpyxl Excel parsing failed: {e}", exc_info=True)
            raise
